import os
import re
import glob
import json
from datetime import datetime
from collections import OrderedDict

import requests
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from scrapy import Spider, Request, signals


class ASMSpider(Spider):
    name = "ASM"
    current_dt = datetime.now().strftime("%d%m%Y%H%M")

    custom_settings = {
        'RETRY_TIMES': 5,
        'RETRY_HTTP_CODES': [500, 502, 503, 504, 400, 403, 404, 408],
        # 'CONCURRENT_REQUESTS': 5,
    }

    headers = {
        'Accept': 'application/json, text/plain, */*',
        'Accept-Language': 'en-US,en;q=0.9',
        'Authorization': '',
        'Connection': 'keep-alive',
        'Content-Type': 'application/json',
        'Origin': 'https://gmppro.asminternational.org',
        'Referer': 'https://gmppro.asminternational.org/',
        'UnitSystem': '0',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
        'ValueReturnMode': 'ActualAndFormattedValue',
    }

    fields = [
        # Remaining Fields
        'Search Material Type', 'Search Material Group', 'Search Material Standard', 'Title',
        'Country/Standard', 'Material Id', 'Producer', 'Material Group', 'Tracker', 'Cross Referencing',

        # New Approach
        # 'Physical Properties', 'Mechanical Properties',
        'Composition Properties'
    ]

    def __init__(self):
        super().__init__()
        # Logs
        os.makedirs('logs', exist_ok=True)
        self.logs_filepath = f'logs/{self.name} International_logs {self.current_dt}.txt'
        self.skipped_urls_filepath = f'logs/{self.name} International_skipped_urls.txt'
        self.script_starting_datetime = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        self.write_logs(f'[INIT] Script started at {self.script_starting_datetime}')

        # self.all_scraped_material_groups = [] # use for writing Scrape Materials groups in json file
        self.current_scraped_mat_group = ''
        self.materials_types = self.get_types()
        self.skipped_urls = list(set(self.read_skipped_urls()))
        self.output_file_path = f'output/{self.name} International Materials Details.xlsx'
        self.previous_scraped_records = self.read_write_xlsx_file(key='previous_records')
        self.standards = self.read_json_files(filepath='input/asm_input/standards.json')[0]
        self.material_groups = self.read_json_files(filepath='input/asm_input/material_groups.json')[0]
        self.all_material_groups = len(self.material_groups)  # for idle function logs
        self.search_json_url = 'https://webapp-tmasm-gateway.azurewebsites.net/horizon/en/total-search/quick-search'

        self.unauth_request = 0
        self.group_found_records = 0
        self.group_scraped_records = 0
        self.all_groups_found_records = 0
        self.all_groups_scraped_records = 0

        self.ready_write_records = []
        self.ready_write_records_counter = 0 # variable declared to write the Excel file when the ten records scraped

    def parse_group(self, response):
        if 'Unauthorized' in response.text:
            self.headers['Authorization'] = ''
            self.current_scraped_mat_group = ''
            self.spider_idle()
            return

        material_type_name = response.meta.get('material_type', {}).get('materialTypeName', '')
        material_group_name = response.meta.get('material_group', {}).get('name', '')
        info = f'Material Type: {material_type_name} & Material Group: {material_group_name}'

        # Handle status code 220
        if response.status == 220:
            self.write_logs(f"{info} returned Status code: {response.status}")
            material_type = response.meta.get('material_type', {})
            material_group_info = response.meta.get('material_group', {})
            for standard in self.standards:
                standard_name = standard.get('standardName', '')
                url_key = f'{material_type_name}_{material_group_name}_{standard_name}'
                if url_key in self.skipped_urls:
                    self.write_logs(f'{info} & Standard: {standard_name} Last Scraped Has no records so its skipped')
                    continue

                response.meta['standard'] = standard
                form_data = self.get_formdata(material_type, material_group_info, key=standard)

                yield Request(url=self.search_json_url, headers=self.headers, dont_filter=True,
                                method='POST', body=json.dumps(form_data),
                                callback=self.parse_group, meta=response.meta)

        # Handle successful responses (status 200)
        if response.status == 200:
            try:
                data_dict = response.json()
            except json.JSONDecodeError as e:
                self.write_logs(f"JSON Decode Error for Type: {type} & Material Group: {material_group_name}-{e}")
                return

            records = data_dict.get('materials', [])
            standard = response.meta.get('standard', {}).get('standardName', '')

            if not records:
                self.write_logs(f"{info} & Standard: {standard} returned No Record")
                if standard:
                    self.write_skipped_urls(f'{material_type_name}_{material_group_name}_{standard}')
                else:
                    self.write_skipped_urls(f'{material_type_name}_{material_group_name}')
                return
            else:
                self.write_logs(f"{info} & Standard: {standard}  found total records: {len(records)}")

            self.all_groups_found_records += len(records)
            self.group_found_records += len(records)
            print('all_groups_found_records :', self.all_groups_found_records)
            print('group_found_records :', self.group_found_records)

            for record in records[:2]: #test
                material_id = record.get('materialId', 0)
                country = record.get('standardName', '')
                title= record.get('designation', '').strip()
                mat_grp = record.get('classification', {})[0].strip()

                if any(tag in str(title) for tag in ['<sub>', '<sup>']):
                    title = self.convert_to_scientific_notation(str(title))
                if any(tag in str(mat_grp) for tag in ['<sub>', '<sup>']):
                    mat_grp = self.convert_to_scientific_notation(str(mat_grp))

                record_key = f'{title}_{material_id}_{mat_grp}_{country}'
                if record_key in self.previous_scraped_records:
                    self.all_groups_scraped_records += 1
                    self.group_scraped_records += 1
                    print(f'Record Already scraped so skipped...')
                    print('Group Items Scraped:', self.group_scraped_records)
                    print('All Groups Records Scraped:', self.all_groups_scraped_records)
                    continue

                if material_id:
                    item = {}
                    material_standard_name = response.meta.get('standard', {}).get('standardName', '')
                    material_standard_country = response.meta.get('standard', {}).get('countryName', '')
                    material_standard = f'Country: {material_standard_country} & Name: {material_standard_name}' if material_standard_name and material_standard_country else ''
                    material_type_name = response.meta.get('material_type', {}).get('materialTypeName', '')
                    material_group_name = response.meta.get('material_group', {}).get('name', '')

                    if any(tag in str(material_type_name) for tag in ['<sub>', '<sup>']):
                        material_type_name = self.convert_to_scientific_notation(str(material_type_name))
                    if any(tag in str(material_group_name) for tag in ['<sub>', '<sup>']):
                        material_group_name = self.convert_to_scientific_notation(str(material_group_name))

                    item['Search Material Type'] =  material_type_name
                    item['Search Material Group'] = material_group_name
                    item['Search Material Standard'] = material_standard
                    item['Title'] = title
                    item['Country/Standard'] = country
                    item['Producer'] = record.get('producers', '')
                    item['Material Id'] = material_id
                    material_grp = '\n '.join(record.get('classification', [])).strip()
                    if any(tag in str(material_grp) for tag in ['<sub>', '<sup>']):
                        material_grp = self.convert_to_scientific_notation(str(material_grp))

                    item['Material Group'] = material_grp
                    group_url = f'https://webapp-tmasm-gateway.azurewebsites.net/horizon/en/materials/{material_id}/tabs/info'
                    response.meta['record'] = record
                    response.meta['item'] = item
                    yield Request(group_url, headers=self.headers, callback=self.parse_detail, meta=response.meta)

    def parse_detail(self, response):
        if 'Unauthorized' in response.text:
            self.headers['Authorization'] =''
            self.current_scraped_mat_group = ''
            self.spider_idle()
            return

        material_type = response.meta.get('material_type', {}).get('materialTypeName', '')
        material_group = response.meta.get('material_group', {}).get('name', '')
        info = f'Material Type: {material_type} & Material Group: {material_group}'
        # record = response.meta.get('record', {})

        if response.status != 200:
            return

        item = response.meta.get('item', {})
        try:
            data_dict = response.json()
        except json.JSONDecodeError as e:
            print('error', e)
            self.write_logs(f'{info}  Record Url: {response.url} Has status code: {response.status}')
            return

        try:
            for key, property_rec in data_dict.items():
                if key == 'mechanicalGroupInfo' and property_rec is not None:
                    # item['Mechanical Properties'] = self.extract_dynamic_fields(property_rec, item, prefix="Mechanical")
                    self.extract_dynamic_fields(property_rec, item, prefix="Mechanical")

                elif key == 'physicalGroupInfo' and property_rec is not None:
                    # item['Physical Properties'] = self.extract_dynamic_fields(property_rec, item, prefix="Physical")
                    self.extract_dynamic_fields(property_rec, item, prefix="Physical")

                elif key== 'trackerGroupInfo' and property_rec is not None:
                    item['Tracker'] = ', '.join(property_rec.get('propertyGroupNames', []))

                elif key== 'crossRefGroupInfo' and property_rec is not None:
                    item['Cross Referencing'] = property_rec.get('total', 0)

        except Exception as e:
            self.write_logs(f'Error Parse Detail Function {info} : Material Name {item['Title']}  URL:{response.url} Error:{e}')

        composition_url = f'https://webapp-tmasm-gateway.azurewebsites.net/horizon/en/materials/{item['Material Id']}/chemical/plus/data/'
        yield Request(url=composition_url, headers=self.headers, callback=self.parse_composition_properties, meta={'item': item, "handle_httpstatus_all": True})

    def parse_composition_properties(self, response):
        item = OrderedDict()
        item.update(response.meta.get('item', {}))
        try:
            data_dict = response.json()
        except json.JSONDecodeError as e:
            self.write_logs(f'Composition Properties error: {e}  URL:{response.url}')
            data_dict=[]

        properties_list = []
        properties_dict = {}
        for record in data_dict:
            min_value = record.get('minValue', '') or ''
            max_value = record.get('maxValue', '') or ''
            approx_value = record.get('approxValue', '') or ''

            try:
                if any(tag in str(max_value) for tag in ['<sub>', '<sup>']):
                    max_value = self.convert_to_scientific_notation(str(max_value))
                if any(tag in str(approx_value) for tag in ['<sub>', '<sup>']):
                    approx_value = self.convert_to_scientific_notation(str(approx_value))

                value = f'{min_value}-{max_value}' if min_value and max_value else min_value or max_value or approx_value
                component = record.get('propertyName', '').strip()
                unit = record.get('unitName', '') or ''
                cas_number = record.get('casNumber', '') or ''
                cas_number = cas_number.replace('-', '') if cas_number else ''

                if any(tag in str(component) for tag in ['<sub>', '<sup>']):
                    component = self.convert_to_scientific_notation(str(component))

                field_name = component

                final_value = f'{value}_{unit}_{cas_number}' if value and unit and cas_number else f'{value}_{unit}'
                if final_value=='_':
                    continue

                properties_list.append(f"{field_name}: {value}")
                # Add to dictionary instead of list
                properties_dict[field_name] = value

            except Exception as e:
                print('Error :', e)

        properties_list = ',\n'.join(properties_list)
        # item['Composition Properties'] = properties_list
        item['Composition Properties'] = properties_dict
        self.group_scraped_records += 1
        self.all_groups_scraped_records += 1
        print('Group Items Scraped:', self.group_scraped_records)
        print('All Groups Records Scraped:', self.all_groups_scraped_records)

        self.read_write_xlsx_file(record=item, key='record')

    def convert_to_scientific_notation(self, text):
        if '<sup>' in text:
            superscript_map = {'0': '\u2070', '1': '\u00b9', '2': '\u00b2', '3': '\u00b3',
                               '4': '\u2074', '5': '\u2075', '6': '\u2076', '7': '\u2077',
                               '8': '\u2078', '9': '\u2079', '-': '\u207b'}  # Include '-' for negative exponents

            # Replace <sup> tags with Unicode superscript characters
            return re.sub(r'<sup>(-?\d+)</sup>', lambda match: ''.join(superscript_map.get(char, char) for char in match.group(1)), text)

        elif '<sub>' in text:
            # This regex finds numbers within <sub> tags and replaces them with subscript Unicode characters
            text = re.sub(r'<sub>(\d+)</sub>',
                          lambda match: ''.join([chr(0x2080 + int(digit)) for digit in match.group(1)]), text)
            return text
        else:
            return text

    def read_json_files(self, filepath):
        files = glob.glob(filepath)

        data = []
        for file in files:
            try:
                with open(file, 'r', encoding='utf-8') as f:
                    file_data = json.load(f)
                    data.append(file_data)

            except json.JSONDecodeError as e:
                print(f"Error decoding JSON in file {file}: {e}")
            except Exception as e:
                print(f"Unexpected error while processing file {file}: {e}")

        data=data

        return data

    def get_material_groups(self):
        res = requests.get(
            'https://webapp-tmasm-gateway.azurewebsites.net/horizon/en/material-groups',
            headers=self.headers,
        )

        if res.status_code == 200:
            try:
                data_dict = res.json()
            except json.JSONDecodeError as e:
                print(f"Error decoding JSON: {e}")
                return

            self.extract_material_groups(data_dict)
            return self.material_groups
        else:
            print(f"Failed to fetch data. Status code: {res.status_code}")

    def extract_material_groups(self, data):
        for item in data:
            children = item.get('children', [])
            if children:
                self.extract_material_groups(children)
            else:
                # Add the current group's id and description to the list
                id = item.get('id', 0)
                desc = item.get('description', '')
                name_value = item.get('name', '')

                if name_value and desc and name_value != desc:
                    name = f'{desc} ({name_value})'
                else:
                    name = name_value or desc

                # name = f'{desc} ({name_value})' if desc else name_value
                self.material_groups.append({
                    'id': id,
                    'name': name,
                    'level': item.get('level', 0),
                })

    def get_formdata(self, material_type,material_group, key):
        json_data = {
            'commonSearchType': 2,
            'searchTerm': '',
            'standardsList': [],
            'materialType': {
                'materialTypeID': material_type['materialTypeID'],
                'materialTypeName': material_type['materialTypeName'],
                'parentID': None,
                'children': [],
            },
            'materialGroups': [
                {
                    'nameAdditionalContent': '',
                    'name': material_group['name'],
                    'id': material_group['id'],
                    'isPolymer': False,
                    'isBiopolymer': False,
                    'show': True,
                    'level': material_group['level'],
                    'expandable': False,
                    'tmMaterialGroupId': None,
                },
            ],
            'showAll': False,
        }
        if key:
            json_data['standardsList'] = [
                {
                    'standardId': key.get('standardId'),
                    'standardName': key.get('standardName', ''),
                    'countryId': key.get('countryId'),
                    'countryName': key.get('countryName', ''),
                }
            ]
        return json_data

    def get_types(self):
        json_data = [
                {
                    "materialTypeID": 1,
                    "materialTypeName": "Bulk Materials",
                     },
                {
                    "materialTypeID": 2,
                    "materialTypeName": "Additive Manufacturing",
                    },
                {
                    "materialTypeID": 3,
                    "materialTypeName": "Adhesives",
                    },
                {
                    "materialTypeID": 4,
                    "materialTypeName": "Coatings",
                },
                {
                    "materialTypeID": 5,
                    "materialTypeName": "Lubricants",
                },
                {
                    "materialTypeID": 6,
                    "materialTypeName": "Substances",
                }
            ]
        return json_data

    def write_logs(self, log_msg):
        with open(self.logs_filepath, mode='a', encoding='utf-8') as logs_file:
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def write_skipped_urls(self, log_msg):
        # Ensure the file exists
        if not os.path.exists(self.skipped_urls_filepath):
            with open(self.skipped_urls_filepath, 'w', encoding='utf-8') as logs_file:
                pass  # Create the file if it doesn't exist

        # Read and write operations ensuring full-line match
        with open(self.skipped_urls_filepath, 'r+', encoding='utf-8') as logs_file:
            # Check if the exact log_msg exists in the file
            if any(line.strip() == log_msg for line in logs_file):
                print(f'{log_msg} already in Skipped URLs. Skipped...')
                return
            # Append the log_msg to the file
            logs_file.write(f'{log_msg}\n')
            print(log_msg)

    def read_skipped_urls(self):
        """Read skipped URLs from the file, if available."""
        try:
            file_path = glob.glob('logs/ASM International_skipped_urls.txt')[0]
            with open(file_path, 'r', encoding='utf-8') as file:
                return [line.strip() for line in file if line.strip()]
        except (IndexError, FileNotFoundError):
            self.write_logs("Skipped URLs file not found.")
        except Exception as e:
            self.write_logs(f"Error reading skipped URLs: {e}")
        return []

    def read_write_xlsx_file(self,record=None, key=None):
        """
            Reads and writes data to/from Excel and JSON files.

            Args:
                self: An instance of the class containing the function.
                record: A dictionary representing a single record to be written.
                key: A string indicating the operation to perform:
                    - 'previous_records': Read existing data from the Excel file.
                    - 'record': Write a new record to the JSON file.
                    - 'close_spider': Write all accumulated records to the Excel file
                                     and delete the JSON file.

            Returns:
                A dictionary of existing records if key is 'previous_records',
                an empty dictionary otherwise.
        """
        json_file = f'logs/{self.name} International Materials Details.json'
        try:
            if key=='previous_records':
                file_exists = os.path.exists(self.output_file_path)
                if not file_exists:
                    return {}

                # Load the workbook (assume the first match)
                workbook = load_workbook(self.output_file_path)
                sheet = workbook.active
                headers = [cell.value for cell in sheet[1]]

                data = {}
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    dict_row = dict(zip(headers, row))
                    title = dict_row.get('Title', '')
                    material_group = dict_row.get('Material Group', '')
                    country= dict_row.get('Country/Standard', '')
                    mat_id= dict_row.get('Material Id', '')
                    key = f"{title}_{mat_id}_{material_group}_{country}"
                    data[key] = dict_row
                return data

            elif key=='record' or key=='close_spider':
                if record:
                    self.ready_write_records.append(record)
                self.ready_write_records_counter += 1

                # Only write to the file when we have 1000 records
                if len(self.ready_write_records) >= 1000 or key=='close_spider':
                    try:
                        #first create json file temporary to enhance the script performance
                        with open(json_file, 'r', encoding='utf-8') as f:
                            existing_data = json.load(f)
                    except FileNotFoundError:
                        existing_data = []
                        for rec_key, rec_value in self.previous_scraped_records.items():
                            if '<sup>' in rec_value:
                                rec_value= self.convert_to_scientific_notation(str(rec_value))
                            existing_data.append(rec_value)

                    # Efficiently merge existing and new records
                    for record in self.ready_write_records:
                        for field in self.fields:
                            record[field] = record.get(field, 'N/A') or 'N/A'
                    existing_data.extend(self.ready_write_records)

                    with open(json_file, 'w', encoding='utf-8') as f:
                        json.dump(existing_data, f, indent=4, ensure_ascii=False)

                    self.ready_write_records = []
                    if key=='close_spider':
                        #first read all records from the json File json_file
                        with open(json_file, 'r', encoding='utf-8') as f:
                            all_records = json.load(f)

                        file_exists = os.path.exists(self.output_file_path)
                        # If the file exists, delete it
                        if file_exists:
                            os.remove(self.output_file_path)

                        # Create a new workbook and add headers
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Materials"
                        ws.append(self.fields)

                        #23-1-25
                        # Write each record to the sheet
                        for record in all_records:
                            # Process the record by ensuring each field has a valid value or 'N/A'
                            processed_record = []
                            for field in self.fields:
                                value = record.get(field, 'N/A')
                                # Handle nested dictionaries by converting them to a string
                                if isinstance(value, dict):
                                    # Convert dictionary to JSON string without escaping non-ASCII characters
                                    value = json.dumps(value, ensure_ascii=False)
                                elif value in (None, '', ' '):  # Replace empty values with 'N/A'
                                    value = 'N/A'
                                processed_record.append(value)

                            # Append the processed record to the sheet
                            ws.append(processed_record)

                        # Save the updated workbook to the file
                        wb.save(self.output_file_path)

                        # #old approach saved data into Excel not in dictionaries
                        # # Write each record to the sheet
                        # for record in all_records:
                        #     # Process the record by ensuring each field has a valid value or 'N/A'
                        #     processed_record = [
                        #         record.get(field, 'N/A') if record.get(field, 'N/A') not in (None, '', ' ') else 'N/A'
                        #         for field in self.fields
                        #     ]
                        #     # Append the processed record to the sheet
                        #     ws.append(processed_record)
                        #
                        # # Save the updated workbook to the file
                        # wb.save(self.output_file_path)
                        # # Process each record from all_records
                        # for record in all_records:
                        #     processed_record = []
                        #     for field in self.fields:
                        #         value = record.get(field, 'N/A')
                        #
                        #         # If the value is a list of dictionaries, convert it to a human-readable string
                        #         if isinstance(value, list):
                        #             value = ",\n ".join([
                        #                     ", ".join(f"{k}: {v}" for k, v in item.items()) if isinstance(item, dict) else str(item)
                        #                     for item in value])
                        #
                        #         # Handle empty or invalid values
                        #         if value in (None, '', ' '):
                        #             value = 'N/A'
                        #
                        #         processed_record.append(value)
                        #
                        #     # Append the processed record to the sheet
                        #     ws.append(processed_record)
                        #
                        # # Save the updated workbook to the file
                        # wb.save(self.output_file_path)
                        os.remove(json_file)

        except Exception as e:
            self.write_logs(f"An error occurred: {e}")
            return {}

    def extract_dynamic_fields(self, property_rec, item, prefix=""):
        """
        Extract all dynamic fields from a given property dictionary and append to item.
        Also add fields dynamically to self.fields for tracking.
        """
        try:
            # properties_list = []
            # Iterate over all properties
            properties = property_rec.get('properties', [])
            for prop in properties:
                field_name = f"{prefix}-{prop.get('name', '').strip()}"
                # field_name = prop.get('name', '').strip()
                min_value = prop.get('valueMin_', '') or prop.get('valueMin', 0.0) or ''
                max_value = prop.get('valueMax_', '') or prop.get('valueMax', 0.0) or ''
                unit = prop.get('unit', '') or ''

                if any(tag in str(unit) for tag in ['<sub>', '<sup>']):
                    unit = self.convert_to_scientific_notation(str(unit))
                if any(tag in str(field_name) for tag in ['<sub>', '<sup>']):
                    field_name = self.convert_to_scientific_notation(str(field_name))

                # Convert min and max values to human-readable format
                if min_value and max_value and unit and min_value == max_value:
                    # value = f"{min_value}-{unit}".strip()
                    value = f"{min_value}".strip()
                elif min_value and max_value:
                    value = f"{min_value} to {max_value}".strip() if unit else f"{min_value}-{max_value}".strip()
                elif min_value:
                    value = f"{min_value}".strip()
                elif max_value:
                    value = f"{max_value}".strip()
                else:
                    value = unit.strip()

                # Special case for text fields like 'Color'
                if field_name.endswith('Color'):
                    value = prop.get('valueTxt', '').strip()
                    field_name = field_name.replace('-', '') if field_name else field_name

                # Fallback to 'valueTxt' if no value is determined
                value = value or prop.get('valueTxt', '') or ''

                if value==' to ':
                    continue

                if any(tag in str(value) for tag in ['<sub>', '<sup>']):
                    value = self.convert_to_scientific_notation(str(value))

                field_name = f'{field_name}-{unit}'

                if field_name not in self.fields:
                    self.fields.append(field_name)

                item[field_name]= value
            #     properties_list.append(f"{field_name}: {value}")
            # properties=  ',\n'.join(properties_list)
            # return properties
            return
        except Exception as e:
            self.write_logs(f"Error extracting dynamic fields: {e}")

    def write_fields_to_file(self):
        """
        Writes all the collected field names in self.fields to a text file.
        Skips fields that already exist in the file.
        """
        try:
            # Ensure the output directory exists
            folder_name = "output"
            os.makedirs(folder_name, exist_ok=True)

            # Define the output file path
            output_file = f'{folder_name}/{self.name}_fields_names.txt'

            # Read existing fields from the file (if it exists)
            existing_fields = set()
            if os.path.exists(output_file):
                with open(output_file, 'r', encoding='utf-8') as f:
                    existing_fields = {line.strip() for line in f}

            # Filter out fields that already exist
            new_fields = [field for field in self.fields if field not in existing_fields]

            # Write only new fields to the file
            if new_fields:
                with open(output_file, 'a', encoding='utf-8') as f:
                    for field in new_fields:
                        f.write(f"{field}\n")
                print(f"Added {len(new_fields)} new fields to {output_file}")
            else:
                print(f"No new fields to add. All fields are already present in {output_file}")

        except Exception as e:
            self.write_logs(f"Error writing fields to file: {e}")

    def close(Spider, reason):
        Spider.write_fields_to_file()
        Spider.read_write_xlsx_file(record=None, key='close_spider')

        # Log overall scraping statistics
        Spider.write_logs(f"\n--- Scraping Summary ---")
        Spider.write_logs(f"Total Products Available on Website: {Spider.all_groups_found_records}")
        Spider.write_logs(f"Total Products Successfully Scraped: {Spider.all_groups_scraped_records}")

        # Log script execution times
        Spider.write_logs(f"\n--- Script Execution Times ---")
        Spider.write_logs(f"Script Start Time: {Spider.script_starting_datetime}")
        Spider.write_logs(f"Script End Time: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        Spider.write_logs(f"Reason for Closure: {reason}")

    @classmethod
    def from_crawler(cls, crawler, *args, **kwargs):
        spider = super(ASMSpider, cls).from_crawler(crawler, *args, **kwargs)
        crawler.signals.connect(spider.spider_idle, signal=signals.spider_idle)
        return spider

    def spider_idle(self):
        if self.headers['Authorization'] == '':
            auth_token = input("Enter the Authorization header value: ").strip()
            auth_token = auth_token.replace("'", "").replace(",", "")
            self.headers['Authorization'] = auth_token

        if self.current_scraped_mat_group:
            self.write_skipped_urls(f'{self.current_scraped_mat_group}')
            self.current_scraped_mat_group = ''

        if self.material_groups:
            # Log progress
            self.write_logs(
                f"\n\n{len(self.material_groups)}/{self.all_material_groups} Material Group Subcategories left to Scrape\n\n")

            self.group_found_records = 0
            self.group_scraped_records = 0

            scraped_material_group = self.material_groups.pop()
            mat_group_name = scraped_material_group.get('name', '')
            if '<sub>' in mat_group_name:
                mat_group_name = self.convert_to_scientific_notation(mat_group_name)

            self.current_scraped_mat_group = mat_group_name
            self.write_logs(f'Material Group: {mat_group_name} Starting for Scraped')
            if self.current_scraped_mat_group in self.skipped_urls:  # if Material Group Already Scraped
                self.write_logs(f'Material Group: {mat_group_name} already scraped last time.')
                self.spider_idle()
                return

            # Track whether at least one request was made
            made_request = False
            # Iterate over material types to generate requests
            for material_type in self.materials_types:
                mat_type_name = material_type.get('materialTypeName', '')
                mat_group_type = f'{mat_type_name}_{mat_group_name}'
                print("material_type", material_type)
                if mat_group_type in self.skipped_urls:
                    self.write_logs(
                        f'Material Type: {mat_type_name} & Material Group: {mat_group_name} Last Scraped Has no records so its skipped')
                    continue

                made_request = True
                form_data = self.get_formdata(material_type, scraped_material_group, key=False)
                self.crawler.engine.crawl(Request(url=self.search_json_url, headers=self.headers, method="POST",
                                                  body=json.dumps(form_data), callback=self.parse_group,
                                                  dont_filter=True,
                                                  meta={"handle_httpstatus_all": True,
                                                        "material_type": material_type,
                                                        "material_group": scraped_material_group}))

            # If no requests were made, move to close the spider
            if not made_request:
                self.spider_idle()

